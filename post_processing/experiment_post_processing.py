import numpy as np
import pandas as pd
import seaborn as sns
import fnmatch
import matplotlib.pyplot as plt
from itertools import product
import openpyxl
import os
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
import math


#%% parameters
# parameters

confidences_list            = [0, 0.2, 0.5, 0.7, 0.9]
pred_steps_list             = [1, 6]
model_type_designator_list  = {"Full" : "multi_topic", "No Topics" : "no_topics", "No Sentiment" : "no_sentiment"}
df_columns                  = ["Time Steps (Mins)", "Confidence"] + list(model_type_designator_list.keys()) + ["Bet Up Every Time", "Bet Down Every Time"]
target_columns_string_dict  = {"bets_proportion" : "bets_with_confidence_proportion_sX_c{}", "precision" : "x_mins_PC_sX_c{}", "score" : "x_mins_score_sX_c{}", "weighted_score" : "x_mins_weighted_sX_c{}"}
up_down_outputs_cols_to_records_cols_dict = {"Bet Up Every Time" : "up", "Bet Down Every Time" : "down"}
DoE_ID_cut_off = 20
optimisation_cut_off = 26

# importing of data
results_csv_file_home_path  = "C:\\Users\\Public\\fabio_uni_work\\Social-Media-and-News-Article-Sentiment-Analysis-for-Stock-Market-Autotrading\\human_readable_results\\240220_export_a"
results_csv_name_list       = ["run32_3.csv", "run32_4.csv", "run32_5.csv", "run32_9.csv", "run32_10.csv", "run32_11.csv"]
#results_csv_file_home_path  = "C:\\Users\\Public\\fabio_uni_work\\Social-Media-and-News-Article-Sentiment-Analysis-for-Stock-Market-Autotrading\\human_readable_results\\240228_export_reduced_dimension"
#results_csv_name_list       = ["run31reducedDesignDimension_noSentiment_5.csv", "run31reducedDesignDimension_noSentiment_11.csv"]


always_up_down_csv_file_path= "C:\\Users\\Public\\fabio_uni_work\\Social-Media-and-News-Article-Sentiment-Analysis-for-Stock-Market-Autotrading\\precalculated_assets\\always_up_down_results\\always_up_down.csv"
df_always_up_down_results   = pd.read_csv(always_up_down_csv_file_path) #FG_fix
df_always_up_down_results.set_index("bet_direction", inplace=True) #FG_fix

# exporting data

outputs_folder_file_path    = "C:\\Users\\Public\\fabio_uni_work\\Social-Media-and-News-Article-Sentiment-Analysis-for-Stock-Market-Autotrading\\human_readable_results\\"
overall_results_file        = "C:\\Users\\Public\\fabio_uni_work\\Social-Media-and-News-Article-Sentiment-Analysis-for-Stock-Market-Autotrading\\human_readable_results\\Final_Project_Excel_Results.xlsx"


# settings

remove_old_financial_scaling_value = True


#%% scraps of code

"""cols_strings_ori = results_csv.columns
pattern = re.compile(r's\d+_')

cols_strings_new = []
for i, input_string in enumerate(cols_strings_ori):
    matches = pattern.findall(input_string)
    replaced_string = pattern.sub('sX_', input_string)
    if not replaced_string in cols_strings_new:
        cols_strings_new = cols_strings_new + [replaced_string]"""
    

#%% methods
print("start")
def to_scientific_notation(number, sf):
    number = float(number)
    format_string = "{{:.{}e}}".format(sf)
    scientific_notation = format_string.format(number)
    return scientific_notation

def return_average_of_con_scope_and_output_combo(df_results, target_output_col, model_type, pred_steps):
    df_results = df_results[df_results["run_name"].str.contains(model_type_designator_list[model_type], case=False)]
    df_results = df_results[df_results["pred_steps"] == pred_steps]
    return df_results[target_output_col].mean()


def return_composite_index_value(df_results, target_output_col, model_type, pred_steps, output_tables_dict_single, con=0, target_columns_string_dict=target_columns_string_dict):
    global COMPOSITE_POWER, action_min_str
    col_str = output_tables_dict_single["col_str"]
    if "{}"in col_str:
        col_str = col_str.format(con)
    
    df_results = df_results[df_results["run_name"].str.contains(model_type_designator_list[model_type], case=False)]
    df_results = df_results[df_results["pred_steps"] == pred_steps]

    
    if action_min_str in output_tables_dict_single["special_actions"]:
        output_numer = df_results["testing_" + col_str] * (df_results["validation_" + col_str] ** - COMPOSITE_POWER)
        output_denom = df_results["validation_" + col_str] ** - COMPOSITE_POWER
    else:
        output_numer = df_results["testing_" + col_str] * (df_results["validation_" + col_str] ** COMPOSITE_POWER)
        output_denom = df_results["validation_" + col_str] ** COMPOSITE_POWER
    return sum(output_numer) / sum(output_denom)





def return_pareto_results_and_ID(df_results, target_output_col, model_type, pred_steps, output_tables_dict_single, target_columns_string_dict=target_columns_string_dict):
    df_results = df_results[df_results["run_name"].str.contains(model_type_designator_list[model_type], case=False)]
    df_results = df_results[df_results["pred_steps"] == pred_steps]
    if action_min_str in output_tables_dict_single["special_actions"]:
        opt_id      = df_results[target_output_col].idxmin()
    else:
        opt_id      = df_results[target_output_col].idxmax()
    opt_value   = df_results.at[opt_id, target_output_col]
        
    return opt_id, opt_value 

def return_col_for_always_up_down_dict(input_str, pred_steps):
    
    input_str = input_str[:input_str.rfind("c")+1] + "0"
    input_str = input_str[:input_str.rfind("sX")+1] + str(int(pred_steps)) + input_str[input_str.rfind("sX")+2:]
    
    return input_str

def return_target_column(output_tables_dict_single, con):
    if "mae" in output_tables_dict_single["col_str"]:
        output = output_tables_dict_single["group"] + "_" + output_tables_dict_single["col_str"]
    else:
        output = str(output_tables_dict_single["group"] + "_" + output_tables_dict_single["col_str"]).format(con)
        
    return output

def return_column_string_2(input_string, con):
    if "{}" in input_string:
        output = input_string.format(con)
    else:
        output = input_string
    return output
        
def comply_with_request_to_drop_fin_scale_value(df, remove_old_financial_scaling_value):
    if remove_old_financial_scaling_value == True and True == False:
        df = df[df['fin_inputs_params_dict_financial_value_scaling'] != 2]
    return df



    

def return_pareto_table_and_max_ids(df_results, output_tables_dict_single, max_ids_and_vals_dict, sf=1):
    global action_no_pre_str, action_pop_ID, action_n_updn, action_ave_str, action_composite_score_str
    
    # the function of the input bools can be found at the large if-else gate
    df_output = pd.DataFrame(columns=df_columns)
    
    for pred_steps in pred_steps_list:
        for con in confidences_list:
            id = len(df_output.index)
            df_output.loc[id, ["Time Steps (Mins)", "Confidence"]] = [pred_steps, con]
            for model_type in model_type_designator_list:
                # this if gate controls what result is populated. See each if statement for more info FG_ACTION_UPDATE NPTE
                if action_no_pre_str in output_tables_dict_single["special_actions"] and action_no_val_str in output_tables_dict_single["special_actions"]:
                    raise ValueError("can't have both no_pre_ID and no_val in special actions")
                target_col = return_target_column(output_tables_dict_single, con)
                if action_composite_score_str in output_tables_dict_single["special_actions"]:
                    val = return_composite_index_value(df_results, target_col, model_type, pred_steps, output_tables_dict_single, con=con)
                    df_output.loc[id, model_type] = val
                elif action_ave_str in output_tables_dict_single["special_actions"]:
                    val = return_average_of_con_scope_and_output_combo(df_results, target_col, model_type, pred_steps)
                    df_output.loc[id, model_type] = val
                elif action_no_pre_str in output_tables_dict_single["special_actions"]:
                    val_id, val = return_pareto_results_and_ID(df_results, target_col, model_type, pred_steps, output_tables_dict_single, output_tables_dict_single)
                    df_output.loc[id, model_type] = val
                    # update new max_ids_and_vals_dict value
                    max_ids_and_vals_dict[pred_steps, con, model_type] = (val_id, val)
                elif not action_no_pre_str in output_tables_dict_single["special_actions"] and not action_no_val_str in output_tables_dict_single["special_actions"]:
                    val = df_results.loc[max_ids_and_vals_dict[pred_steps, con, model_type][0], return_column_string_2(target_col, con)]
                    df_output.loc[id, model_type] = val
                elif action_no_val_str in output_tables_dict_single["special_actions"]:
                    df_output.loc[id, model_type] = max_ids_and_vals_dict[pred_steps, con, model_type][0]
                
                # populate always up/down results
                if not action_no_up_down_str in output_tables_dict_single["special_actions"]:
                    #records_col = target_columns_string_dict["target_output_name"].format(pred_steps, 0)
                    target_updown_col =  return_col_for_always_up_down_dict(target_col, pred_steps)
                    for output_col in up_down_outputs_cols_to_records_cols_dict:
                        records_row = up_down_outputs_cols_to_records_cols_dict[output_col]
                        val = df_always_up_down_results.loc[records_row, target_updown_col[target_updown_col.find("_")+1:]]
                        df_output.loc[id, output_col] = val
                
    df_output = df_output.apply(pd.to_numeric, errors='coerce')

    
    return df_output, max_ids_and_vals_dict
    

def update_excel(df, start_cell, excel_file_path, pop_table_labels=False, table_name="Blank Name"):
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        raise ValueError("the formating from this file is needed")
    sheet = workbook.active
    for row_num, row in enumerate(df.values):
        for col_num, value in enumerate(row):
            cell = sheet.cell(row=row_num + sheet[start_cell].row, column=col_num + sheet[start_cell].column)
            cell.value = value
    if pop_table_labels==True:
        table_name = table_name.replace(" (", "\n(")
        cell = sheet.cell(row=sheet[start_cell].row - 2, column=sheet[start_cell].column)
        cell.value = table_name
        for i_col in range(len(df.columns)):
            if math.isnan(df.iloc[0, i_col]) == False:
                cell = sheet.cell(row=sheet[start_cell].row - 1, column=sheet[start_cell].column + i_col)
                cell.value = df.columns[i_col]
    
    workbook.save(excel_file_path)


def convert_table_time_value_from_timesteps_to_mins(df, mins_per_timestep=5):
    df["Time Steps (Mins)"] = mins_per_timestep * df["Time Steps (Mins)"]
    return df

def trim_and_new_line_long_variable_names(input):
    # this method returns a reduced version of the string or strings entered, optionally with new line characters.
    if isinstance(df_results_filtered.columns, pd.core.indexes.base.Index):
        input = list(input)
    elif not isinstance(input, list):
        input = [input]
    
    stopwords = ["params", "dict", "dicts", "inputs", "hyper", "x_mins", "sX"]
    replacement_pairs = [("validation", "val"), ("testing", "test"), ("temporal", "temp"), ("reporting", "report")]
    for i in range(len(input)):
        var = input[i]
        #remove stopwords
        for word in stopwords:
            var = var.replace(word, "")
        for pair in replacement_pairs:
            var = var.replace(pair[0], pair[1])
        while "__" in var:
            var = var.replace("__", "_")
        input[i] = var
    return input

#%% main line

action_no_pre_str       = "dont_populate_according_to_previous_max_id"
action_no_val_str       = "populate_ID_value_not_val"
action_min_str          = "minimise_value_not_maximise"
action_no_up_down_str   = "dont_populate_up_down_values"
action_ave_str          = "report_averages"
action_composite_score_str  = "report_composite_score"
COMPOSITE_POWER         = 3

weighted_score_col_str  = "x_mins_weighted_sX_c{}"
absolute_score_col_str  = "x_mins_score_sX_c{}"
bets_proportion_col_str = "bets_with_confidence_proportion_sX_c{}"
precision_col_str       = "x_mins_PC_sX_c{}"

output_tables_dict = {
    
    "Optimum Validation MAE Score" : 
        {"col_str":"mae",
        "group" : "validation", "top_left_cell" : "B4",
        "special_actions" : [action_no_pre_str, action_min_str, action_no_up_down_str]},
    "Validation Weighted Score (for Lowest Validation MAE Score Design)":
        {"col_str":weighted_score_col_str,
        "group" : "validation", "top_left_cell" : "J4",
        "special_actions" : []},
    "Validation Absolute Score (for Lowest Validation MAE Score Design)":
        {"col_str":absolute_score_col_str,
        "group" : "validation", "top_left_cell" : "R4",
        "special_actions" : []},
    "Validation Proportion of Correct Positions (for Lowest Validation MAE Score Design)":
        {"col_str":precision_col_str,
        "group" : "validation", "top_left_cell" : "Z4",
        "special_actions" : []},
    "Global Design IDs (for Lowest Validation MAE Score Design)":
        {"col_str":"mae",
        "group" : "testing", "top_left_cell" : "AP4",
        "special_actions" : [action_no_val_str, action_no_up_down_str]},

    "Testing MAE Score (for Lowest Validation MAE Score Design)":
        {"col_str":"mae",
        "group" : "testing", "top_left_cell" : "B31",
        "special_actions" : [action_no_up_down_str]},
    "Testing Weighted Score (for Lowest Validation MAE Score Design)":
        {"col_str":weighted_score_col_str,
        "group" : "testing", "top_left_cell" : "J31",
        "special_actions" : []},
    "Testing Absolute Score (for Lowest Validation MAE Score Design)":
        {"col_str":absolute_score_col_str,
        "group" : "testing", "top_left_cell" : "R31",
        "special_actions" : []},
    "Testing Proportion of Correct Positions (for Lowest Validation MAE Score Design)":
        {"col_str":precision_col_str,
        "group" : "testing", "top_left_cell" : "Z31",
        "special_actions" : []},

    ## score optimised results

    "Optimum Weighted Confidence":
        {"col_str":weighted_score_col_str,
        "group" : "validation", "top_left_cell" : "J55",
        "special_actions" : [action_no_pre_str]},
    "Validation MAE score (for Highest Validation Weighted Confidence Design)": 
        {"col_str":"mae",
        "group" : "validation", "top_left_cell" : "B55",
        "special_actions" : [action_no_up_down_str]},
    "Validation Absolute Score (for Highest Validation Weighted Confidence Design)":
        {"col_str":absolute_score_col_str,
        "group" : "validation", "top_left_cell" : "R55",
        "special_actions" : []},
    "Validation Proportion of Correct Positions (for Highest Validation Weighted Confidence Design)":
        {"col_str":precision_col_str,
        "group" : "validation", "top_left_cell" : "Z55",
        "special_actions" : []},
    "Global Design IDs (for Highest Validation Weighted Confidence Design)":
        {"col_str":"mae",
        "group" : "testing", "top_left_cell" : "AP55",
        "special_actions" : [action_no_val_str, action_no_up_down_str]},
    "Testing MAE Score (for Highest Validation Weighted Confidence Design)":
        {"col_str":"mae",
        "group" : "testing", "top_left_cell" : "B82",
        "special_actions" : [action_no_up_down_str]},
    "Testing Weighted Score (for Highest Validation Weighted Confidence Design)":
        {"col_str":weighted_score_col_str,
        "group" : "testing", "top_left_cell" : "J82",
        "special_actions" : []},
    "Testing Absolute Score (for Highest Validation Weighted Confidence Design)":
        {"col_str":absolute_score_col_str,
        "group" : "testing", "top_left_cell" : "R82",
        "special_actions" : []},
    "Testing Proportion of Correct Positions (for Highest Validation Weighted Confidence Design)":
        {"col_str":precision_col_str,
        "group" : "testing", "top_left_cell" : "Z82",
        "special_actions" : []},
    
    ## declare averages

    "Average Validation MAE Score" : 
        {"col_str" : "mae",
        "group" : "validation", "top_left_cell" : "B120",
        "special_actions" : [action_ave_str, action_no_up_down_str]},
    "Average Validation Weighted Score":
        {"col_str" : weighted_score_col_str,
        "group" : "validation", "top_left_cell" : "J120",
        "special_actions" : [action_ave_str]},
    "Average Validation Absolute Score":
        {"col_str" : absolute_score_col_str,
        "group" : "validation", "top_left_cell" : "R120",
        "special_actions" : [action_ave_str]},
    "Average Validation Proportion of Correct Positions":
        {"col_str" : precision_col_str,
        "group" : "validation", "top_left_cell" : "Z120",
        "special_actions" : [action_ave_str]},

    "Average Testing MAE Score":
        {"col_str" : "mae",
        "group" : "testing", "top_left_cell" : "B143",
        "special_actions" : [action_ave_str, action_no_up_down_str]},
    "Average Testing Weighted Score":
        {"col_str" : weighted_score_col_str,
        "group" : "testing", "top_left_cell" : "J143",
        "special_actions" : [action_ave_str]},
    "Average Testing Absolute Score":
        {"col_str" : absolute_score_col_str,
        "group" : "testing", "top_left_cell" : "R143",
        "special_actions" : [action_ave_str]},
    "Average Testing Proportion of Correct Positions":
        {"col_str" : precision_col_str,
        "group" : "testing", "top_left_cell" : "Z143",
        "special_actions" : [action_ave_str]},


    "Average Training MAE Score":
        {"col_str" : "mae",
        "group" : "training", "top_left_cell" : "z107",
        "special_actions" : [action_ave_str, action_no_up_down_str]},



    ## declare composites
    "MAE Score Composite Index":
        {"col_str" : "mae",
        "group" : "testing", "top_left_cell" : "B166",
        "special_actions" : [action_composite_score_str, action_min_str, action_no_up_down_str]},
    "Weighted Score Composite Index":
        {"col_str" : weighted_score_col_str,
        "group" : "testing", "top_left_cell" : "J166",
        "special_actions" : [action_composite_score_str]},
    "Absolute Score Composite Index":
        {"col_str" : absolute_score_col_str,
        "group" : "testing", "top_left_cell" : "R166",
        "special_actions" : [action_composite_score_str]},
    "Proportion of Correct Positions Composite Index":
        {"col_str" : precision_col_str,
        "group" : "testing", "top_left_cell" : "Z166",
        "special_actions" : [action_composite_score_str]}
    }




#%%


# collate results


df_results = pd.DataFrame()
for results_file in results_csv_name_list:
    df_temp = pd.read_csv(os.path.join(results_csv_file_home_path, results_file))
    df_temp = df_temp[df_temp["local_ID"]<=optimisation_cut_off]
    df_results = pd.concat([df_results, df_temp], axis=0)
df_results = df_results.dropna(subset=['experiment_timestamp'])
df_results = df_results.reset_index(drop=True)

df_results = comply_with_request_to_drop_fin_scale_value(df_results, remove_old_financial_scaling_value)
df_results.iloc[:,0] = df_results.index
df_results.to_csv(outputs_folder_file_path + "\\" + "combined_csv" + ".csv")
df_results_T = df_results.T
df_results_T.to_csv(outputs_folder_file_path + "\\" + "combined_csv_T" + ".csv")

max_ids_and_vals_dict       = dict()
final_output_tables_list    = dict()

for name in output_tables_dict:
    df, max_ids_and_vals_dict = return_pareto_table_and_max_ids(df_results, output_tables_dict[name], max_ids_and_vals_dict, output_tables_dict)
    df = convert_table_time_value_from_timesteps_to_mins(df)
    df.to_csv(outputs_folder_file_path + "\\" + name + ".csv")
    #save_table(df, generate_image_name(target_output_name), folder, colour_bool)
    update_excel(df, output_tables_dict[name]["top_left_cell"], overall_results_file, pop_table_labels=False, table_name=name)
    
    final_output_tables_list[name] = df
print("complete tables")


#%% make graphics

# validation/testing, mae and weighted score
substring_categories = ['multi_topic', 'no_topics', 'no_sentiment']
pred_steps_list     = pred_steps_list

for pred_steps in pred_steps_list:
    df_results_filtered = df_results[df_results["pred_steps"]==pred_steps]
    df_results_filtered.loc[:,'category'] = df_results_filtered['run_name'].apply(lambda x: next((cat for cat in substring_categories if cat in x), 'Other'))
    df_results_filtered = df_results_filtered[["local_ID", 'category', "validation_mae", "testing_mae", "validation_x_mins_weighted_sX_c0.9", "testing_x_mins_weighted_sX_c0.9"]]
    df_results_filtered[["local_ID", "validation_mae", "testing_mae", "validation_x_mins_weighted_sX_c0.9", "testing_x_mins_weighted_sX_c0.9"]] = df_results_filtered[["local_ID", "validation_mae", "testing_mae", "validation_x_mins_weighted_sX_c0.9", "testing_x_mins_weighted_sX_c0.9"]].astype(float)

    # Set the color palette for the categories
    palette = {'multi_topic': 'blue', 'no_topics': 'green', 'no_sentiment': 'orange', 'Other': 'gray'}

    # Create a scatter plot using Seaborn
    plt.figure(figsize=(10, 6))
    pairplot = sns.pairplot(df_results_filtered, hue='category', palette=palette, plot_kws=dict(marker="+", linewidth=1))
    plt.suptitle(f'pairplot {5*pred_steps} mins')
    # Add legend
    pairplot.add_legend(title='Categories')


    # Create a folder to save individual subplots


    # Save the overall pairplot
    pairplot.savefig(os.path.join(outputs_folder_file_path, f'pairplot {5*pred_steps} mins.png'))


    # Save each subplot separately
    outputs_folder = outputs_folder_file_path + "\\pairplot_subplots"
    os.makedirs(outputs_folder, exist_ok=True)

    for i in df_results_filtered.columns[2:]:
        for j in df_results_filtered.columns[2:]:
            for cat in [None] + substring_categories:
                for results_subset in ["Full", "DoE_only", "optim_only"]:
                    if cat == None:
                        df_results_filtered_2 = df_results_filtered
                        naming_suffix = f'_{pred_steps*5}mins'
                    else:
                        df_results_filtered_2 = df_results_filtered[df_results_filtered["category"] == cat]
                        naming_suffix = f'_{pred_steps*5}mins_{cat}_only'
                    plt.figure(figsize=(8, 6))
                    if results_subset == "Full":
                        df_results_filtered_3 = df_results_filtered_2
                    elif results_subset == "DoE_only":
                        df_results_filtered_3 = df_results_filtered_2[df_results_filtered_2["local_ID"]<=DoE_ID_cut_off]
                        naming_suffix += "_DoEonly"
                    elif results_subset == "optim_only":
                        df_results_filtered_3 = df_results_filtered_2[df_results_filtered_2["local_ID"]>DoE_ID_cut_off]
                        naming_suffix += "_optim_only"
                    else:
                        raise ValueError("error found")

                    df_results_filtered_3 = df_results_filtered_3.drop("local_ID", axis=1)
                    if i == j:
                        sub_pairplot = sns.displot(df_results_filtered_3, x=i, hue='category', palette=palette, kind="kde", fill=True)#, plot_kws=dict(marker="+", linewidth=1))
                        plt.suptitle("{} density{}".format(i, naming_suffix))
                        plt.savefig(os.path.join(outputs_folder, f'densityplot_{i}_vs_{j}{naming_suffix}.png'))
                    else:
                        sub_pairplot = sns.scatterplot(df_results_filtered_3, x=i, y=j, hue='category', palette=palette, style="category")#"marker="+")#, plot_kws=dict(marker="+", linewidth=1))
                        temp_title = trim_and_new_line_long_variable_names([f'Scatter Plot of {i} vs {j}{naming_suffix}'])
                        sub_pairplot.set_title(temp_title)
                        #standarised lims between the same charts
                        x0, x1, y0, y1 = min(df_results_filtered[i]), max(df_results_filtered[i]), min(df_results_filtered[j]), max(df_results_filtered[j])
                        xr, yr = x1-x0, y1-y0
                        x0, x1, y0, y1 = x0-xr*0.05, x1+xr*0.05, y0-yr*0.05, y1+yr*0.05
                        sub_pairplot.set(xlim=(x0,x1), ylim=(y0,y1))  # Adjust the limits as needed
                        sub_pairplot.legend(loc='upper left', bbox_to_anchor=(1, 1))
                        plt.savefig(os.path.join(outputs_folder, f'scatterplot_{i}_vs_{j}{naming_suffix}.png'), bbox_inches='tight')
                    plt.close()

                    # do the parallel axis for the category/pred steps combo
                    # https://plotly.com/python/parallel-coordinates-plot/

                    fig = go.Figure()
                    fig = px.parallel_coordinates(df_results_filtered_3, color="testing_x_mins_weighted_sX_c0.9", color_continuous_scale=px.colors.diverging.RdBu)
                    fig.write_html("C:/Users/Public/fabio_uni_work/Social-Media-and-News-Article-Sentiment-Analysis-for-Stock-Market-Autotrading/human_readable_results/parallel_axis_{}{}.html".format(cat, naming_suffix))

# create design parameter parallel plots

design_params_columns = ["local_ID", "run_name", "pred_steps"] + fnmatch.filter(df_results.columns, "*param*") + ["validation_x_mins_weighted_sX_c0.9", "testing_x_mins_weighted_sX_c0.9"]
design_table_folder_name = "parallel_axis_design_tables"
df_results_filtered = df_results[design_params_columns]
df_results_filtered['category'] = df_results_filtered['run_name'].apply(lambda x: next((cat for cat in substring_categories if cat in x), 'Other'))

for pred_steps in pred_steps_list:
    for cat in [None] + substring_categories:
        print("dddddd")
        df_designs_filtered_2 = df_results_filtered.copy()
        df_designs_filtered_2 = df_designs_filtered_2[df_designs_filtered_2["pred_steps"]==pred_steps]
        if cat == None:
            naming_suffix = f'_{pred_steps*5}mins'
        else:
            df_designs_filtered_2 = df_designs_filtered_2[df_designs_filtered_2["category"] == cat]
            naming_suffix = f'_{pred_steps*5}mins_{cat}_only'
        df_designs_filtered_2.columns = trim_and_new_line_long_variable_names(df_designs_filtered_2.columns)
        df_designs_filtered_2 = df_designs_filtered_2.drop("category", axis=1)
        fig = go.Figure()
        fig = px.parallel_coordinates(df_designs_filtered_2) #fig = px.parallel_coordinates(df_designs_filtered_2, color=df_designs_filtered_2.columns[-1], color_continuous_scale=px.colors.diverging.RdBu)
        fig.write_html(os.path.join(outputs_folder_file_path, "parallel_axis_design_tables", f'para_axis {naming_suffix}.html'))
        
        
        # # create a pairplot
        # palette = {'multi_topic': 'blue', 'no_topics': 'green', 'no_sentiment': 'orange', 'Other': 'gray'}
        # # Create a scatter plot using Seaborn
        # plt.figure(figsize=(10, 6))
        # pairplot = sns.pairplot(df_results_filtered, hue='category', palette=palette, plot_kws=dict(marker="+", linewidth=1))
        # plt.suptitle(f'pairplot {5*pred_steps} mins')
        # # Add legend
        # pairplot.add_legend(title='Categories')

            
print("complete all plots")














# %%
