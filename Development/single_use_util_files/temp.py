import openpyxl
import pandas as pd

# Load the Excel file
excel_file_path = 'your_excel_file.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

# Select the specific sheet you want to modify (e.g., the first sheet)
sheet = workbook.active

# Load the DataFrame with the values you want to update

df = pd.DataFrame(data)

def update_excel(df, start_cell, excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    for row_num, row in enumerate(df.values):
        for col_num, value in enumerate(row):
            cell = sheet.cell(row=row_num + sheet[start_cell].row, column=col_num + sheet[start_cell].column)
            cell.value = value
    workbook.save(excel_file_path)
    

# Specify the range of cells to update (assuming the DataFrame size matches the range)
start_cell = 'A1'
end_cell = 'C3'

# Update the values in the Excel file with the DataFrame values
for row_num, row in enumerate(df.values):
    for col_num, value in enumerate(row):
        cell = sheet.cell(row=row_num + sheet[start_cell].row, column=col_num + sheet[start_cell].column)
        cell.value = value

# Save the modified Excel file
workbook.save('modified_excel_file.xlsx')
